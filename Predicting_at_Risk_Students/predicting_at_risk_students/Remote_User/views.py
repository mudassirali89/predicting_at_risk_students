from django.db.models import Count
from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404
import datetime
import openpyxl

from sklearn.feature_extraction.text import CountVectorizer

import pandas as pd

from sklearn.feature_extraction.text import CountVectorizer
from sklearn.metrics import accuracy_score, confusion_matrix, classification_report
from sklearn.metrics import accuracy_score
from sklearn.metrics import f1_score
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import VotingClassifier
from sklearn.ensemble import RandomForestClassifier

# Create your views here.
from Remote_User.models import ClientRegister_Model,student_marks_model,student_risk_prediction_model,detection_ratio_model

def login(request):


    if request.method == "POST" and 'submit1' in request.POST:

        username = request.POST.get('username')
        password = request.POST.get('password')
        try:
            enter = ClientRegister_Model.objects.get(username=username,password=password)
            request.session["userid"] = enter.id

            return redirect('Add_DataSet_Details')
        except:
            pass

    return render(request,'RUser/login.html')

def Add_DataSet_Details(request):
    if "GET" == request.method:
        return render(request, 'RUser/Add_DataSet_Details.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(excel_file)
        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)
        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)
        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)
        # reading a cell
        print(worksheet["A1"].value)
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)
            student_marks_model.objects.all().delete()
            student_risk_prediction_model.objects.all().delete()
    for r in range(1, active_sheet.max_row+1):
        student_marks_model.objects.create(
            regno= active_sheet.cell(r, 1).value,
            names= active_sheet.cell(r, 2).value,
            sem1= active_sheet.cell(r, 3).value,
            sem2= active_sheet.cell(r, 4).value,
            sem3= active_sheet.cell(r, 5).value
        )

    return render(request, 'RUser/Add_DataSet_Details.html', {"excel_data": excel_data})


def Register1(request):

    if request.method == "POST":
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        phoneno = request.POST.get('phoneno')
        country = request.POST.get('country')
        state = request.POST.get('state')
        city = request.POST.get('city')
        ClientRegister_Model.objects.create(username=username, email=email, password=password, phoneno=phoneno,
                                            country=country, state=state, city=city)

        return render(request, 'RUser/Register1.html')
    else:
        return render(request,'RUser/Register1.html')

def ViewYourProfile(request):
    userid = request.session['userid']
    obj = ClientRegister_Model.objects.get(id= userid)
    return render(request,'RUser/ViewYourProfile.html',{'object':obj})


def Predict_Earliest_Possible_Prediction_DataSets(request):
    se=''
    if request.method == "POST":
        kword = request.POST.get('keyword')
        if request.method == "POST":
            kword = request.POST.get('keyword')
            print(kword)

            df = pd.read_csv('Student_DataSets.csv', encoding='latin-1')
            df
            df.columns
            df.rename(columns={'names': 'sname', 'sem3': 's3per'}, inplace=True)

            def apply_results(results):
                if (results < 70):
                    return 0  # At Risk
                else:
                    return 1  # No Risk

            df['results'] = df['s3per'].apply(apply_results)

            X = df['sname']
            y = df['results']

            cv = CountVectorizer(lowercase=False, strip_accents='unicode', ngram_range=(1, 1))

            x = cv.fit_transform(X)
            models = []
            from sklearn.model_selection import train_test_split
            X_train, X_test, y_train, y_test = train_test_split(x, y, test_size=0.20)
            X_train.shape, X_test.shape, y_train.shape

            models = []
            from sklearn.model_selection import train_test_split
            X_train, X_test, y_train, y_test = train_test_split(x, y, test_size=0.20)
            X_train.shape, X_test.shape, y_train.shape

            print("Naive Bayes")

            from sklearn.naive_bayes import MultinomialNB

            NB = MultinomialNB()
            NB.fit(X_train, y_train)
            predict_nb = NB.predict(X_test)
            naivebayes = accuracy_score(y_test, predict_nb) * 100
            print("ACCURACY")
            print(naivebayes)
            print("CLASSIFICATION REPORT")
            print(classification_report(y_test, predict_nb))
            print("CONFUSION MATRIX")
            print(confusion_matrix(y_test, predict_nb))
            models.append(('naive_bayes', NB))

            # SVM Model
            print("SVM")
            from sklearn import svm

            lin_clf = svm.LinearSVC()
            lin_clf.fit(X_train, y_train)
            predict_svm = lin_clf.predict(X_test)
            svm_acc = accuracy_score(y_test, predict_svm) * 100
            print("ACCURACY")
            print(svm_acc)
            print("CLASSIFICATION REPORT")
            print(classification_report(y_test, predict_svm))
            print("CONFUSION MATRIX")
            print(confusion_matrix(y_test, predict_svm))
            models.append(('SVM', lin_clf))

            print("Logistic Regression")

            from sklearn.linear_model import LogisticRegression

            reg = LogisticRegression(random_state=0, solver='lbfgs').fit(X_train, y_train)
            y_pred = reg.predict(X_test)
            print("ACCURACY")
            print(accuracy_score(y_test, y_pred) * 100)
            print("CLASSIFICATION REPORT")
            print(classification_report(y_test, y_pred))
            print("CONFUSION MATRIX")
            print(confusion_matrix(y_test, y_pred))
            models.append(('LogisticRegression', reg))

            print("Decision Tree Classifier")
            dtc = DecisionTreeClassifier()
            dtc.fit(X_train, y_train)
            dtcpredict = dtc.predict(X_test)
            print("ACCURACY")
            print(accuracy_score(y_test, dtcpredict) * 100)
            print("CLASSIFICATION REPORT")
            print(classification_report(y_test, dtcpredict))
            print("CONFUSION MATRIX")
            print(confusion_matrix(y_test, dtcpredict))
            models.append(('DecisionTreeClassifier', dtc))

            print("Gradient Boosting Classifier")
            from sklearn.ensemble import GradientBoostingClassifier
            clf = GradientBoostingClassifier(n_estimators=100, learning_rate=1.0, max_depth=1, random_state=0).fit(
                X_train,
                y_train)
            clfpredict = clf.predict(X_test)
            print("ACCURACY")
            print(accuracy_score(y_test, clfpredict) * 100)
            print("CLASSIFICATION REPORT")
            print(classification_report(y_test, clfpredict))
            print("CONFUSION MATRIX")
            print(confusion_matrix(y_test, clfpredict))
            models.append(('GradientBoostingClassifier', clf))

            classifier = VotingClassifier(models)
            classifier.fit(X_train, y_train)
            y_pred = classifier.predict(X_test)

            kword1 = [kword]
            vector1 = cv.transform(kword1).toarray()
            predict_text = classifier.predict(vector1)

            pred = str(predict_text).replace("[", "")
            pred1 = str(pred.replace("]", ""))

            prediction = int(pred1)

            if prediction == 0:
                val = 'Student At Risk'
            elif prediction == 1:
                val = 'Student At No Risk'

            print(prediction)
            print(val)

        return render(request, 'RUser/Predict_Earliest_Possible_Prediction_DataSets.html',{'objs': val})
    return render(request, 'RUser/Predict_Earliest_Possible_Prediction_DataSets.html')


def ratings(request,pk):
    vott1, vott, neg = 0, 0, 0
    objs = student_marks_model.objects.get(id=pk)
    unid = objs.id
    vot_count = student_marks_model.objects.all().filter(id=unid)
    for t in vot_count:
        vott = t.ratings
        vott1 = vott + 1
        obj = get_object_or_404(student_marks_model, id=unid)
        obj.ratings = vott1
        obj.save(update_fields=["ratings"])
        return redirect('Add_DataSet_Details')

    return render(request,'RUser/ratings.html',{'objs':vott1})
